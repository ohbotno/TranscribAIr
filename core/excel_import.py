"""
Excel rubric import functionality.
Supports auto-detection of simple and analytic rubric formats.
"""
from pathlib import Path
from typing import Optional, List, Tuple
from dataclasses import dataclass

from .rubric import Rubric, RubricCriterion, PerformanceLevel


@dataclass
class ParsedRubricData:
    """Intermediate representation of parsed Excel data."""
    rubric_name: str
    rubric_description: str
    criteria_data: List[dict]
    is_analytic: bool  # True if has performance levels
    performance_level_names: List[str]  # Column headers for performance levels


class ExcelRubricImporter:
    """Imports rubrics from Excel files with auto-format detection."""

    # Common performance level keywords for detection (case-insensitive)
    PERFORMANCE_LEVEL_KEYWORDS = [
        'poor', 'below', 'adequate', 'good', 'very good', 'excellent',
        'exceptional', 'satisfactory', 'unsatisfactory', 'developing',
        'proficient', 'advanced', 'beginning', 'emerging', 'meets', 'exceeds'
    ]

    @staticmethod
    def is_available() -> bool:
        """Check if openpyxl is available."""
        try:
            import openpyxl
            return True
        except ImportError:
            return False

    @staticmethod
    def _normalize_header(header: str) -> str:
        """Normalize column header for comparison."""
        if header is None:
            return ""
        return str(header).strip().lower()

    @staticmethod
    def _is_performance_level_column(header: str) -> bool:
        """Check if column header looks like a performance level."""
        normalized = ExcelRubricImporter._normalize_header(header)
        return any(keyword in normalized for keyword in ExcelRubricImporter.PERFORMANCE_LEVEL_KEYWORDS)

    @staticmethod
    def _extract_percentage_weight(value) -> float:
        """Extract numeric weight from various formats (5%, 0.05, 5, etc.)."""
        if value is None:
            return 1.0

        # Convert to string and clean
        str_value = str(value).strip().replace('%', '')

        try:
            numeric_value = float(str_value)
            # If value is >= 1, assume it's a percentage (5 means 5%, not 500%)
            if numeric_value >= 1:
                return numeric_value / 100.0
            # If < 1, assume it's already a decimal
            return numeric_value
        except ValueError:
            return 1.0

    @classmethod
    def parse_excel(cls, file_path: Path) -> Optional[ParsedRubricData]:
        """
        Parse Excel file and detect rubric format.

        Expected formats:
        1. Simple rubric: Criterion Name | Weight | Description
        2. Analytic rubric: Criterion Name | Weight | Poor | Below | Adequate | Good | Very Good | Exceptional

        The rubric name can be in:
        - First cell (A1)
        - A cell labeled "Rubric Name:" or similar
        - Filename as fallback
        """
        # Dynamic import to allow installation at runtime
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Extract rubric name and description
            rubric_name, rubric_description = cls._extract_metadata(ws, file_path)

            # Find header row (first row with column headers)
            header_row, headers = cls._find_header_row(ws)
            if not header_row or not headers:
                raise ValueError("Could not find column headers in Excel file")

            # Detect rubric format
            is_analytic, performance_level_columns = cls._detect_format(headers)

            # Parse criteria data
            criteria_data = cls._parse_criteria(ws, header_row, headers, is_analytic, performance_level_columns)

            return ParsedRubricData(
                rubric_name=rubric_name,
                rubric_description=rubric_description,
                criteria_data=criteria_data,
                is_analytic=is_analytic,
                performance_level_names=[headers[idx] for idx in performance_level_columns]
            )

        except Exception as e:
            print(f"Error parsing Excel file: {e}")
            return None

    @classmethod
    def _extract_metadata(cls, worksheet, file_path: Path) -> Tuple[str, str]:
        """Extract rubric name and description from worksheet."""
        rubric_name = file_path.stem  # Default to filename
        rubric_description = ""

        # Check first few cells for rubric name
        for row in range(1, 5):
            for col in range(1, 3):
                cell_value = worksheet.cell(row, col).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # Look for metadata labels
                    normalized = cls._normalize_header(cell_str)
                    if 'rubric name' in normalized or 'rubric:' in normalized:
                        # Name should be in next cell or after colon
                        if ':' in cell_str:
                            rubric_name = cell_str.split(':', 1)[1].strip()
                        else:
                            next_cell = worksheet.cell(row, col + 1).value
                            if next_cell:
                                rubric_name = str(next_cell).strip()
                    elif 'description' in normalized:
                        if ':' in cell_str:
                            rubric_description = cell_str.split(':', 1)[1].strip()
                        else:
                            next_cell = worksheet.cell(row, col + 1).value
                            if next_cell:
                                rubric_description = str(next_cell).strip()

        return rubric_name, rubric_description

    @classmethod
    def _find_header_row(cls, worksheet) -> Tuple[Optional[int], Optional[List[str]]]:
        """Find the row containing column headers."""
        # Look for a row with multiple non-empty cells that look like headers
        for row_idx in range(1, min(20, worksheet.max_row + 1)):
            row_values = [worksheet.cell(row_idx, col).value for col in range(1, worksheet.max_column + 1)]
            non_empty = [v for v in row_values if v is not None and str(v).strip()]

            # Need at least 2 columns (name + something else)
            if len(non_empty) >= 2:
                # Check if this looks like a header row (contains common keywords)
                row_text = ' '.join([str(v).lower() for v in non_empty])
                if any(keyword in row_text for keyword in ['criterion', 'name', 'weight', '%', 'description', 'poor', 'good']):
                    # Found header row
                    headers = [str(v).strip() if v is not None else "" for v in row_values]
                    return row_idx, headers

        return None, None

    @classmethod
    def _detect_format(cls, headers: List[str]) -> Tuple[bool, List[int]]:
        """
        Detect if rubric is analytic (has performance levels).

        Returns:
            (is_analytic, performance_level_column_indices)
        """
        performance_level_cols = []

        for idx, header in enumerate(headers):
            if cls._is_performance_level_column(header):
                performance_level_cols.append(idx)

        # If we have 3+ performance level columns, it's likely analytic
        is_analytic = len(performance_level_cols) >= 3

        return is_analytic, performance_level_cols

    @classmethod
    def _parse_criteria(cls, worksheet, header_row: int, headers: List[str],
                       is_analytic: bool, perf_level_cols: List[int]) -> List[dict]:
        """Parse criteria rows from the worksheet."""
        criteria_data = []

        # Identify key columns
        name_col = cls._find_column_index(headers, ['criterion', 'name', 'criteria'])
        weight_col = cls._find_column_index(headers, ['weight', '%', 'percent'])
        description_col = cls._find_column_index(headers, ['description', 'desc']) if not is_analytic else None

        if name_col is None:
            # Default to first column if not found
            name_col = 0

        # Parse data rows (after header)
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            name = worksheet.cell(row_idx, name_col + 1).value
            if not name or not str(name).strip():
                continue  # Skip empty rows

            criterion = {
                'name': str(name).strip(),
                'weight': 1.0,
                'description': '',
                'performance_levels': []
            }

            # Extract weight
            if weight_col is not None:
                weight_value = worksheet.cell(row_idx, weight_col + 1).value
                criterion['weight'] = cls._extract_percentage_weight(weight_value)

            if is_analytic:
                # Parse performance levels
                for pl_col_idx in perf_level_cols:
                    pl_value = worksheet.cell(row_idx, pl_col_idx + 1).value
                    if pl_value and str(pl_value).strip():
                        # Try to extract score range from header
                        header = headers[pl_col_idx]
                        score_range = cls._extract_score_range(header)

                        criterion['performance_levels'].append({
                            'name': headers[pl_col_idx],
                            'score_range': score_range,
                            'description': str(pl_value).strip()
                        })
            else:
                # Parse simple description
                if description_col is not None:
                    desc_value = worksheet.cell(row_idx, description_col + 1).value
                    if desc_value:
                        criterion['description'] = str(desc_value).strip()

            criteria_data.append(criterion)

        return criteria_data

    @staticmethod
    def _find_column_index(headers: List[str], keywords: List[str]) -> Optional[int]:
        """Find column index by matching keywords in header."""
        for idx, header in enumerate(headers):
            normalized = ExcelRubricImporter._normalize_header(header)
            if any(keyword in normalized for keyword in keywords):
                return idx
        return None

    @staticmethod
    def _extract_score_range(header: str) -> str:
        """Extract score range from header like 'Poor <40%' or 'Good 60-70%'."""
        import re

        # Look for patterns like <40%, 40-50%, >80%, etc.
        patterns = [
            r'[<>]?\s*\d+%',  # <40%, 40%, >80%
            r'\d+\s*-\s*\d+%',  # 40-50%
        ]

        for pattern in patterns:
            match = re.search(pattern, header)
            if match:
                return match.group().strip()

        # If no range found, return empty string
        return ""

    @classmethod
    def import_from_excel(cls, file_path: Path) -> Optional[Rubric]:
        """
        Import a complete rubric from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Rubric object or None if import fails
        """
        parsed_data = cls.parse_excel(file_path)
        if not parsed_data:
            return None

        # Convert parsed data to Rubric object
        criteria = []
        for crit_data in parsed_data.criteria_data:
            if parsed_data.is_analytic and crit_data['performance_levels']:
                # Create criterion with performance levels
                perf_levels = [
                    PerformanceLevel(
                        name=pl['name'],
                        score_range=pl['score_range'],
                        description=pl['description']
                    )
                    for pl in crit_data['performance_levels']
                ]
                criterion = RubricCriterion(
                    name=crit_data['name'],
                    weight=crit_data['weight'],
                    performance_levels=perf_levels
                )
            else:
                # Simple criterion
                criterion = RubricCriterion(
                    name=crit_data['name'],
                    description=crit_data['description'],
                    weight=crit_data['weight']
                )

            criteria.append(criterion)

        return Rubric(
            name=parsed_data.rubric_name,
            description=parsed_data.rubric_description,
            criteria=criteria
        )
